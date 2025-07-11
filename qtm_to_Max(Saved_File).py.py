import time
import csv
import asyncio
from datetime import datetime

import qtm_rt
from qtm_rt.packet import QRTComponentType
from pythonosc.udp_client import SimpleUDPClient

import numpy as np

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# --- QTM configuration ---
QTM_HOST = '127.0.0.1'  # QTM server address
OSC_HOST = '139.19.40.134'
UDP_port = 12345
client = SimpleUDPClient(OSC_HOST, UDP_port)
osc_address = '/qtm'

# --- Duration for logging (in seconds) ---
DURATION_SECONDS = 1000

# --- Output file ---
from pathlib import Path
downloads_folder = Path(r"C:\Users\aykumar\Downloads")
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = downloads_folder / f"touch_log_{timestamp}.xlsx"
print(f"📁 Output will be saved to: {OUTPUT_FILE}")


log_rows = []

latest_qtm = None

def rect_point_to_local_xy(corners, point):
    """
    Convert 3D point on rectangle to local 2D coordinates in rectangle frame.
    """
    corners = np.array(corners)
    point = np.array(point)
    center = np.mean(corners, axis=0)
    u = corners[1] - corners[0]
    v = corners[3] - corners[0]
    u_norm = u / np.linalg.norm(u)
    v_norm = v / np.linalg.norm(v)
    vec = point - center
    x_local = np.dot(vec, u_norm)
    y_local = np.dot(vec, v_norm)
    return x_local, y_local

def distance_point_to_plane(point, plane_point, plane_normal):
    """
    Computes signed distance from `point` to a plane.
    """
    point = np.array(point)
    plane_point = np.array(plane_point)
    plane_normal = np.array(plane_normal)
    plane_normal = plane_normal / np.linalg.norm(plane_normal)
    return np.dot(point - plane_point, plane_normal)

def handle_qtm_data(packet):
    global latest_qtm

    try:
        frame = packet.framenumber
        latest_qtm = []

        print(f"\nFramenumber: {frame}")
        header, markers = packet.get_3d_markers()

        labels = [
            "Pen - 1", "Pen - 2", "Pen - 3", "Pen - 4",
            "Screen - 1", "Screen - 2", "Screen - 3", "Screen - 4"
        ]

        for i, marker in enumerate(markers):
            label = labels[i] if i < len(labels) else f"Marker {i+1}"
            print(f"{label}: X={marker.x:.2f}, Y={marker.y:.2f}, Z={marker.z:.2f}")
            latest_qtm.append([marker.x, marker.y, marker.z])

        if len(latest_qtm) >= 8:
            screen_corners = latest_qtm[4:8]
            pen_tip = latest_qtm[3]

            # 1. Compute plane normal from screen corners
            v1 = np.array(screen_corners[1]) - np.array(screen_corners[0])
            v2 = np.array(screen_corners[3]) - np.array(screen_corners[0])
            normal = np.cross(v1, v2)
            normal /= np.linalg.norm(normal)

            # 2. Compute perpendicular distance to screen plane
            plane_point = screen_corners[0]
            dist = abs(distance_point_to_plane(pen_tip, plane_point, normal))

            print(f"Pen distance to screen: {dist:.2f} mm")
            status = "Not Touching"
            x_local = y_local = 'NaN'

            # 3. Check if pen is "touching"
            if dist < 6.0:  # threshold in mm
                x_local, y_local = rect_point_to_local_xy(screen_corners, pen_tip)
                    # Estimate screen width and height from corners
                width = np.linalg.norm(np.array(screen_corners[1]) - np.array(screen_corners[0]))
                height = np.linalg.norm(np.array(screen_corners[3]) - np.array(screen_corners[0]))
                half_width = width / 2
                half_height = height / 2

                if -half_width <= x_local <= half_width and -half_height <= y_local <= half_height:
                      print(f"✅ Pen Tip TOUCHING screen at local: x={x_local:.2f}, y={y_local:.2f}")
                      status = "Touching (Green)"
                else:
                    print(f"⚠️ Pen is near the screen but **outside bounds**, local: x={x_local:.2f}, y={y_local:.2f}")
                    status = "Outside Bounds (Red)"
            else:
                print("🛑 Pen is not touching the screen.")

            # ✅ Log row for Excel
            log_rows.append([
                frame,
                pen_tip[0], pen_tip[1], pen_tip[2],
                dist,
                x_local if isinstance(x_local, (int, float)) else 'NaN',
                y_local if isinstance(y_local, (int, float)) else 'NaN',
                status
            ])

    except Exception as e:
        print(f"bug {e}")
        latest_qtm = ['NaN', 'NaN', 'NaN', 'NaN']




async def start_qtm_listener():
    conn = await qtm_rt.connect(QTM_HOST)
    await conn.stream_frames(components=['3d', '6d'], on_packet=handle_qtm_data)

async def wait_for_valid_qtm():
    while True:
        await asyncio.sleep(0.01)
        if latest_qtm and isinstance(latest_qtm[0], float):
            break

# --- Main Async Task ---
async def main():
    global log_rows
    asyncio.ensure_future(start_qtm_listener())
    await wait_for_valid_qtm()
    start_time = time.time()

    try:
        while (time.time() - start_time) < DURATION_SECONDS:
            await asyncio.sleep(0.01)
    finally:
        # Save to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Touch Log"
        headers = ['Frame', 'Pen X', 'Pen Y', 'Pen Z', 'Distance to Plane (mm)', 'Local X', 'Local Y', 'Touch Status']
        ws.append(headers)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in log_rows:
            ws.append(row)
            last_row = ws.max_row
            status_cell = ws.cell(row=last_row, column=8)
            if "Green" in row[7]:
                status_cell.fill = green_fill
            elif "Red" in row[7]:
                status_cell.fill = red_fill

        print(f"✅ Data saved to: {OUTPUT_FILE}")
        wb.save(OUTPUT_FILE)


# --- Entry Point ---
if __name__ == '__main__':
    time.sleep(2)
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n⛔ Interrupted by user. Saving data before exiting...")
        # Save to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Touch Log"
        headers = ['Frame', 'Pen X', 'Pen Y', 'Pen Z', 'Distance to Plane (mm)', 'Local X', 'Local Y', 'Touch Status']
        ws.append(headers)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in log_rows:
            ws.append(row)
            last_row = ws.max_row
            status_cell = ws.cell(row=last_row, column=8)
            if "Green" in row[7]:
                status_cell.fill = green_fill
            elif "Red" in row[7]:
                status_cell.fill = red_fill

        wb.save(OUTPUT_FILE)
        print(f"✅ Data saved to: {OUTPUT_FILE}")

