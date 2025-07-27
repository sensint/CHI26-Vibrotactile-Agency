import os
import time
import threading
import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import Workbook
import openpyxl
from screeninfo import get_monitors
import ctypes
from pathlib import Path
import subprocess
import json

SERIAL_PORT = 'COM7'
BAUD_RATE = 115200
W_VALUES = [90, 80, 70, 60, 50]
D_VALUES = [90, 160, 280, 480, 800]

BOUNDS = {'1': [(- (D_VALUES[0]/2 + W_VALUES[0]), - (D_VALUES[0]/2)),(D_VALUES[0]/2, D_VALUES[0]/2 + W_VALUES[0])],
          '2': [(- (D_VALUES[1]/2 + W_VALUES[1]), - (D_VALUES[1]/2)),(D_VALUES[1]/2, D_VALUES[1]/2 + W_VALUES[1])],
          '3': [(- (D_VALUES[2]/2 + W_VALUES[2]), - (D_VALUES[2]/2)),(D_VALUES[2]/2, D_VALUES[2]/2 + W_VALUES[2])],
          '4': [(- (D_VALUES[3]/2 + W_VALUES[3]), - (D_VALUES[3]/2)),(D_VALUES[3]/2, D_VALUES[3]/2 + W_VALUES[3])],
          '5': [(- (D_VALUES[4]/2 + W_VALUES[4]), - (D_VALUES[4]/2)),(D_VALUES[4]/2, D_VALUES[4]/2 + W_VALUES[4])]} 

TOTAL_TRIALS = 20
participant_name = ""
participant_id = ""
vibro_feedback = False
trial_count = {}
difficulty = 1
clicks = 0
target_side = 1
previous_rects = []
data = []
participant_folder = ""
current_rect = (0, 0, 0, 0)
start_time = 0
CANVAS_WIDTH = 0
CANVAS_HEIGHT = 0
experiment_finished = False

script_dir = Path(__file__).parent
session_file = script_dir / "current_session_path.txt"

monitors = get_monitors()
selected_monitor = monitors[1] if len(monitors) > 1 else monitors[0]
root = tk.Tk()
root.withdraw()
experiment_window = tk.Toplevel()
experiment_window.title("Fitts' Law Experiment")
screen_x = selected_monitor.x
screen_y = selected_monitor.y
screen_width = selected_monitor.width
screen_height = selected_monitor.height
experiment_window.geometry(f"{screen_width}x{screen_height}+{screen_x}+{screen_y}")
experiment_window.update_idletasks()
hwnd = ctypes.windll.user32.GetForegroundWindow()
ctypes.windll.user32.MoveWindow(hwnd, screen_x, screen_y, screen_width, screen_height, True)
CANVAS_WIDTH = screen_width
CANVAS_HEIGHT = screen_height
canvas = tk.Canvas(experiment_window, width=CANVAS_WIDTH, height=CANVAS_HEIGHT, bg="white")
canvas.pack()
canvas.pack_forget()
qtm_proc = None

import serial
try:
    ser = serial.Serial(SERIAL_PORT, 9600, timeout=0.1)
except:
    ser = None
    print("Warning: Serial port not available.")

def listen_serial():
    while True:
        if ser and ser.in_waiting:
            line = ser.readline().decode('utf-8').strip()
            if line == '1':
                experiment_window.event_generate('<<SerialClick>>', when='tail')

def draw_rectangle():
    global current_rect, start_time
    canvas.delete("all")
    for rect in previous_rects:
        canvas.create_rectangle(rect[0], rect[1], rect[0] + rect[2], rect[1] + rect[3], fill="lightgrey")
    rect_width = W_VALUES[difficulty - 1]
    distance = D_VALUES[difficulty - 1]
    rect_height = CANVAS_HEIGHT
    center_x = CANVAS_WIDTH / 2
    rect_x = center_x + (target_side * distance) - (rect_width / 2)
    rect_y = 0
    canvas.create_rectangle(rect_x, rect_y, rect_x + rect_width, rect_y + rect_height, fill="blue")
    current_rect = (rect_x, rect_y, rect_width, rect_height)
    start_time = time.time()

def get_latest_clicked_coordinates():
    locals_x = []
    locals_y = []
    insides = []
    try:
        files = sorted(Path(participant_folder).glob("clicked_log_*.xlsx"), reverse=True)
        if not files:
            return None, None, None

        wb = openpyxl.load_workbook(files[0])
        ws = wb.active
        screen_width_mm, screen_height_mm = get_screen_dimensions_mm()

        for i, row in enumerate(list(ws.iter_rows(min_row=2, values_only=True))):
            if row[-1] != 1 or row[5] == 'NaN' or row[6] == 'NaN':
                continue

            clicked_x_mm = float(row[5])
            clicked_y_mm = float(row[6])
            locals_x.append(clicked_x_mm)
            locals_y.append(clicked_y_mm)

            mm_per_px = screen_width_mm / CANVAS_WIDTH
            if i % 2 == 0:
                gt_x1_px, gt_x2_px = BOUNDS["1"][0]
            else:
                gt_x1_px, gt_x2_px = BOUNDS["1"][1]

            gt_x1_mm = gt_x1_px * mm_per_px
            gt_x2_mm = gt_x2_px * mm_per_px

            in_x = gt_x1_mm <= clicked_x_mm <= gt_x2_mm
            insides.append(in_x)

        return locals_x, locals_y, insides

    except Exception as e:
        print(f"Error reading clicked coordinates: {e}")
    return None, None, None

def get_screen_dimensions_mm():
    try:
        files = sorted(Path(participant_folder).glob("clicked_log_*.xlsx"), reverse=True)
        if not files:
            return 346.00, 194.57

        wb = openpyxl.load_workbook(files[0])
        ws = wb.active

        x_vals = [abs(float(r[5])) for r in ws.iter_rows(min_row=2, values_only=True) if r[5] != 'NaN']
        y_vals = [abs(float(r[6])) for r in ws.iter_rows(min_row=2, values_only=True) if r[6] != 'NaN']

        width_mm = max(x_vals) * 2 if x_vals else 346.00
        height_mm = max(y_vals) * 2 if y_vals else 194.57

        return width_mm, height_mm

    except Exception as e:
        print(f"Error getting screen dimensions: {e}")
        return 346.00, 194.57

def handle_serial_click(event=None):
    global clicks, target_side, current_rect, experiment_finished
    if experiment_finished:
        return
    click_time = time.time()
    mt = (click_time - start_time) * 1000
    error = 0
    speed = D_VALUES[difficulty - 1] / mt if mt > 0 else 0
    throughput = difficulty / (mt / 1000) if mt > 0 else 0
    data.append({
    'MT': mt,
    'speed': speed,
    'error': 0,  # default, will correct later
    'throughput': throughput,
    'side': target_side,  # -1 for left, +1 for right
    'LocalX': None        # placeholder
})

    previous_rects.append(current_rect)
    clicks += 1
    target_side *= -1
    if clicks >= TOTAL_TRIALS:
        experiment_finished = True
        messagebox.showinfo("Done", "Experiment complete. Processing results...", parent=experiment_window)
        experiment_window.unbind('<<SerialClick>>')
        end_trial()
    else:
        draw_rectangle()


def save_data_and_finish():
    global experiment_finished
    avg_mt = sum(d['MT'] for d in data) / TOTAL_TRIALS
    avg_speed = sum(d['speed'] for d in data) / TOTAL_TRIALS
    avg_tp = sum(d['throughput'] for d in data) / TOTAL_TRIALS

    locals_x, locals_y, insides = get_latest_clicked_coordinates()

    # ✅ Correct each error in data[] based on QTM-inside status
    if insides:
        for i, inside in enumerate(insides):
            if i < len(data):
                data[i]['error'] = 0 if inside else 1

    # ✅ Compute error rate from corrected data
    error_values = [d['error'] for d in data]
    avg_error = (sum(error_values) / TOTAL_TRIALS if error_values else 0)

    # ✅ Save results
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Headers with LocalX
    headers = ['MT', 'speed', 'error', 'throughput', 'LocalX']
    ws.append(headers)

    # Add each row of trial data with LocalX
    for i, row in enumerate(data):
        local_x_val = locals_x[i] if locals_x and i < len(locals_x) else ''
        ws.append([row.get('MT', ''), row.get('speed', ''), row.get('error', ''), row.get('throughput', ''), local_x_val])

    # ✅ Add final average row (last row)
    ws.append(['AVG', avg_speed, '', avg_tp, ''])

    # Save Excel file
    vibro_label = "VibroOn" if vibro_feedback else "VibroOff"
    filename = f"{participant_name}_{participant_id}_{vibro_label}_ID{difficulty}_trial{trial_count[participant_id][difficulty]}.xlsx"
    filepath = os.path.join(participant_folder, filename)
    wb.save(filepath)

    time.sleep(5)

    # ✅ Show summary popup
    summary = f"Avg MT: {avg_mt:.2f} ms\nAvg Speed: {avg_speed:.2f} px/ms\nAvg Throughput: {avg_tp:.2f} bit/s\nAvg Error: {avg_error*100:.2f}%"
    messagebox.showinfo("Experiment Finished", f"{summary}", parent=experiment_window)
    experiment_window.quit()


def end_trial():
    save_data_and_finish()

def begin_trial():
    global difficulty, clicks, data, previous_rects
    difficulty = 1
    clicks = 0
    data.clear()
    previous_rects.clear()
    trial_count.setdefault(participant_id, {}).setdefault(difficulty, 0)
    trial_count[participant_id][difficulty] += 1
    canvas.pack()
    draw_rectangle()

def start_experiment():
    global participant_name, participant_id, vibro_feedback, participant_folder
    participant_name = simpledialog.askstring("Participant", "Enter Name:", parent=experiment_window)
    participant_id = simpledialog.askstring("Participant", "Enter ID:", parent=experiment_window)
    vibro_feedback = messagebox.askyesno("Vibrotactile", "Enable vibrotactile feedback?", parent=experiment_window)

    base_dir = r"C:\\Users\\aykumar\\Desktop\\Results"
    participant_folder = os.path.join(base_dir, f"{participant_name}_{participant_id}")
    os.makedirs(participant_folder, exist_ok=True)

    session_file.write_text(participant_folder)

    qtm_proc = subprocess.Popen(["python", "qtm.py"], cwd=script_dir)
    print(f"🟢 QTM logger started with PID {qtm_proc.pid}")
    time.sleep(2)
    begin_trial()

btn_start = tk.Button(experiment_window, text="Start Experiment", command=start_experiment, font=("Arial", 20))
btn_start.pack(pady=10)

if ser:
    threading.Thread(target=listen_serial, daemon=True).start()

experiment_window.bind('<<SerialClick>>', handle_serial_click)
experiment_window.mainloop()
