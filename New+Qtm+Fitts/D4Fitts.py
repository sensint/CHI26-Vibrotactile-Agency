# Full experiment script for Difficulty 4
# W = 60 px, D = 480 px
# Rectangle bounds converted to mm for LocalX error detection

import os
import time
import threading
import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import Workbook, load_workbook
from screeninfo import get_monitors
import ctypes
from pathlib import Path
import subprocess
import serial

SERIAL_PORT = 'COM7'
W_VALUES = [60]
D_VALUES = [480]
TOTAL_TRIALS = 20
participant_name = ""
participant_id = ""
vibro_feedback = False
trial_count = {}
difficulty = 1
clicks = 0
target_side = 1
previous_rects = []
data = []
target_sides = []
participant_folder = ""
current_rect = (0, 0, 0, 0)
start_time = 0
CANVAS_WIDTH = 0
CANVAS_HEIGHT = 0
experiment_finished = False

script_dir = Path(__file__).parent
session_file = script_dir / "current_session_path.txt"

monitors = get_monitors()
selected_monitor = monitors[1] if len(monitors) > 1 else monitors[0]
root = tk.Tk()
root.withdraw()
experiment_window = tk.Toplevel()
screen_x = selected_monitor.x
screen_y = selected_monitor.y
screen_width = selected_monitor.width
screen_height = selected_monitor.height
experiment_window.geometry(f"{screen_width}x{screen_height}+{screen_x}+{screen_y}")
experiment_window.update_idletasks()
hwnd = ctypes.windll.user32.GetForegroundWindow()
ctypes.windll.user32.MoveWindow(hwnd, screen_x, screen_y, screen_width, screen_height, True)
CANVAS_WIDTH = screen_width
CANVAS_HEIGHT = screen_height
canvas = tk.Canvas(experiment_window, width=CANVAS_WIDTH, height=CANVAS_HEIGHT, bg="white")
canvas.pack()
canvas.pack_forget()

try:
    ser = serial.Serial(SERIAL_PORT, 9600, timeout=0.1)
except:
    ser = None
    print("Warning: Serial port not available.")

def listen_serial():
    while True:
        if ser and ser.in_waiting:
            line = ser.readline().decode('utf-8').strip()
            if line == '1':
                experiment_window.event_generate('<<SerialClick>>', when='tail')

def draw_rectangle():
    global current_rect, start_time
    canvas.delete("all")
    for rect in previous_rects:
        canvas.create_rectangle(rect[0], rect[1], rect[0]+rect[2], rect[1]+rect[3], fill="lightgrey")
    rect_width = W_VALUES[0]
    distance = D_VALUES[0]
    rect_height = CANVAS_HEIGHT
    center_x = CANVAS_WIDTH / 2
    rect_x = center_x + (target_side * distance) - (rect_width / 2)
    rect_y = 0
    canvas.create_rectangle(rect_x, rect_y, rect_x + rect_width, rect_y + rect_height, fill="blue")
    current_rect = (rect_x, rect_y, rect_width, rect_height)
    start_time = time.time()

def get_screen_dimensions_mm():
    try:
        files = sorted(Path(participant_folder).glob("clicked_log_*.xlsx"), reverse=True)
        if not files:
            return 346.00, 194.57
        wb = load_workbook(files[0])
        ws = wb.active
        x_vals = [abs(float(r[5])) for r in ws.iter_rows(min_row=2, values_only=True) if r[5] != 'NaN']
        y_vals = [abs(float(r[6])) for r in ws.iter_rows(min_row=2, values_only=True) if r[6] != 'NaN']
        return max(x_vals)*2 if x_vals else 346.0, max(y_vals)*2 if y_vals else 194.57
    except:
        return 346.0, 194.57

# Manually derived bounds in mm for W=60 and D=480
# Formula: left = -(D/2 + W), right = -(D/2), and similarly on the right
# So pixel range = [-300, -240] and [240, 300], scaled to mm using 0.1133 (based on 1920px -> 346mm)
# => bounds: [-33.99 mm, -27.19 mm] and [27.19 mm, 33.99 mm]
def is_inside_rectangle(localx):
        return (-91.90 <= localx <= -81.09) or (81.09 <= localx <= 91.90)

def update_fitts_file_with_localx_and_errors(fitts_path, clicked_path, canvas_width, d_value, w_value):
    try:
        wb_f = load_workbook(fitts_path)
        ws_f = wb_f.active

        wb_c = load_workbook(clicked_path)
        ws_c = wb_c.active

        all_rows = list(ws_c.iter_rows(min_row=2, values_only=True))
        valid_rows = [r for r in all_rows if r[-1] == 1 and r[5] not in [None, 'NaN']]
        localxs = [float(r[5]) for r in valid_rows]

        error_col = []

        for i in range(min(len(localxs), TOTAL_TRIALS)):
            localx = localxs[i]
            error = 0 if is_inside_rectangle(localx) else 1
            error_col.append(error)
            ws_f.cell(row=i + 2, column=4, value=error)
            ws_f.cell(row=i + 2, column=5, value=localx)

        for i in range(len(localxs), TOTAL_TRIALS):
            ws_f.cell(row=i + 2, column=4, value=1)
            ws_f.cell(row=i + 2, column=5, value="NaN")
            error_col.append(1)

        ws_f.append([
            "=AVERAGE(A2:A21)",
            "=AVERAGE(B2:B21)",
            "=AVERAGE(C2:C21)",
            "=AVERAGE(D2:D21)",
            ""
        ])

        wb_f.save(fitts_path)
        return sum(error_col) / len(error_col) if error_col else 0

    except Exception as e:
        print("Update failed:", e)
        return 0

def handle_serial_click(event=None):
    global clicks, target_side, current_rect, experiment_finished
    if experiment_finished:
        return
    click_time = time.time()
    mt = (click_time - start_time) * 1000
    error = 0
    speed = D_VALUES[0] / mt if mt > 0 else 0
    throughput = 4 / (mt / 1000) if mt > 0 else 0
    data.append({'MT': mt, 'speed': speed, 'error': error, 'throughput': throughput})
    target_sides.append(target_side)
    previous_rects.append(current_rect)
    clicks += 1
    target_side *= -1
    if clicks >= TOTAL_TRIALS:
        experiment_finished = True
        messagebox.showinfo("Done", "Experiment complete. Processing results...", parent=experiment_window)
        experiment_window.unbind('<<SerialClick>>')
        end_trial()
    else:
        draw_rectangle()

def save_data_and_finish():
    global experiment_finished
    avg_mt = sum(d['MT'] for d in data) / TOTAL_TRIALS
    avg_speed = sum(d['speed'] for d in data) / TOTAL_TRIALS
    avg_tp = sum(d['throughput'] for d in data) / TOTAL_TRIALS

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ['MT', 'speed', 'throughput', 'error', 'LocalX']
    ws.append(headers)
    for row in data:
        ws.append([row.get('MT', ''), row.get('speed', ''), row.get('throughput', ''), row.get('error', ''), ""])
    vibro_label = "VibroOn" if vibro_feedback else "VibroOff"
    filename = f"{participant_name}_{participant_id}_{vibro_label}_ID4_trial{trial_count[participant_id][difficulty]}.xlsx"
    filepath = os.path.join(participant_folder, filename)
    wb.save(filepath)

    time.sleep(5)

    clicked_logs = sorted(Path(participant_folder).glob("clicked_log_*.xlsx"), reverse=True)
    avg_error = 0
    if clicked_logs:
        avg_error = update_fitts_file_with_localx_and_errors(filepath, str(clicked_logs[0]), CANVAS_WIDTH, D_VALUES[0], W_VALUES[0])

    summary = f"Avg MT: {avg_mt:.2f} ms\nAvg Speed: {avg_speed:.2f} px/ms\nAvg Throughput: {avg_tp:.2f} bit/s\nAvg Error: {avg_error:.2%}"
    messagebox.showinfo("Experiment Finished", summary, parent=experiment_window)
    experiment_window.quit()

def end_trial():
    save_data_and_finish()

def begin_trial():
    global difficulty, clicks, data, previous_rects, target_sides
    difficulty = 1
    clicks = 0
    data.clear()
    target_sides.clear()
    previous_rects.clear()
    trial_count.setdefault(participant_id, {}).setdefault(difficulty, 0)
    trial_count[participant_id][difficulty] += 1
    canvas.pack()
    draw_rectangle()

def start_experiment():
    global participant_name, participant_id, vibro_feedback, participant_folder
    participant_name = simpledialog.askstring("Participant", "Enter Name:", parent=experiment_window)
    participant_id = simpledialog.askstring("Participant", "Enter ID:", parent=experiment_window)
    vibro_feedback = messagebox.askyesno("Vibrotactile", "Enable vibrotactile feedback?", parent=experiment_window)
    base_dir = r"C:\\Users\\aykumar\\Desktop\\Results"
    participant_folder = os.path.join(base_dir, f"{participant_name}_{participant_id}")
    os.makedirs(participant_folder, exist_ok=True)
    session_file.write_text(participant_folder)
    subprocess.Popen(["python", "qtm.py"], cwd=script_dir)
    time.sleep(2)
    begin_trial()

btn_start = tk.Button(experiment_window, text="Start Experiment", command=start_experiment, font=("Arial", 20))
btn_start.pack(pady=10)

if ser:
    threading.Thread(target=listen_serial, daemon=True).start()

experiment_window.bind('<<SerialClick>>', handle_serial_click)
experiment_window.mainloop()
