import time
import asyncio
import serial
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import qtm_rt
from qtm_rt.packet import QRTComponentType

import sys
from threading import Thread
from datetime import datetime

from pythonosc.udp_client import SimpleUDPClient

# SERIAL_PORT = 'COM3'
SERIAL_PORT = 'COM4'
BAUD_RATE = 115200
QTM_HOST = '139.19.40.134'
OSC_HOST = '139.19.40.35'
UDP_port = 12345
client = SimpleUDPClient(OSC_HOST, UDP_port)
osc_address = '/qtm'

streaming_enabled = True

session_file = Path(__file__).parent / "current_session_path.txt"
# start_signal = Path(__file__).parent / "start_qtm.flag"
# stop_signal = Path(__file__).parent / "stop_qtm.flag"

# Wait for session folder
for _ in range(60):
    if session_file.exists():
        break
    time.sleep(1)
else:
    print("❌ Session path not found.")
    sys.exit(1)

folder_path = session_file.read_text().strip()
log_dir = Path(folder_path)
log_dir.mkdir(parents=True, exist_ok=True)

# print(f"📡 Waiting for trial start signal...")

log_rows = []
clicked_frames = set()
latest_frame = None

def rect_point_to_local_xy(corners, point):
    corners = np.array(corners)
    point = np.array(point)
    center = np.mean(corners, axis=0)
    u = corners[1] - corners[0]
    v = corners[3] - corners[0]
    u_norm = u / np.linalg.norm(u)
    v_norm = v / np.linalg.norm(v)
    vec = point - center
    x_local = np.dot(vec, v_norm)
    y_local = np.dot(vec, u_norm)
    return x_local, y_local

def distance_point_to_plane(point, plane_point, plane_normal):
    point = np.array(point)
    plane_point = np.array(plane_point)
    plane_normal = np.array(plane_normal)
    plane_normal = plane_normal / np.linalg.norm(plane_normal)
    return np.dot(point - plane_point, plane_normal)

def handle_qtm_data(packet):
    global latest_frame, streaming_enabled

    try:
        frame = packet.framenumber
        latest_frame = frame

        header, markers = packet.get_3d_markers()
        marker_xyz = [[m.x, m.y, m.z] for m in markers]
        if len(marker_xyz) >= 8:
            screen_corners = marker_xyz[5:9]
            pen_tip = marker_xyz[4]

            

            v1 = np.array(screen_corners[1]) - np.array(screen_corners[0])
            v2 = np.array(screen_corners[3]) - np.array(screen_corners[0])
            normal = np.cross(v1, v2)
            normal /= np.linalg.norm(normal)

            dist = abs(distance_point_to_plane(pen_tip, screen_corners[0], normal))
            status = "Not Touching"
            x_local = y_local = 'NaN'

            if dist < 8.0:
                x_local, y_local = rect_point_to_local_xy(screen_corners, pen_tip)
                
                if streaming_enabled:
                    client.send_message(osc_address, [round(x_local,1), round(y_local,1)])

                width = np.linalg.norm(np.array(screen_corners[1]) - np.array(screen_corners[0]))
                height = np.linalg.norm(np.array(screen_corners[3]) - np.array(screen_corners[0]))
                if -width/2 <= x_local <= width/2 and -height/2 <= y_local <= height/2:
                    status = "Touching (Green)"
                else:
                    status = "Outside Bounds (Red)"

            log_rows.append([
                frame, pen_tip[0], pen_tip[1], pen_tip[2],
                dist, x_local, y_local, status
            ])
    except Exception as e:
        print(f"❌ QTM error: {e}")

def listen_serial():
    try:
        print(f"🔗 Attempting to connect to serial port {SERIAL_PORT} at {BAUD_RATE} baud...")
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=0.1)
        print("✅ Serial connection established successfully!")
        while True:
            line = ser.readline().decode().strip()
            if line == '1' and latest_frame is not None:
                clicked_frames.add(latest_frame)
                print(f"🔘 Button clicked at frame: {latest_frame}")
    except serial.SerialException as e:
        print(f"❌ Serial port error: {e}")
        print(f"   Could not open {SERIAL_PORT}. Check if:")
        print("   1. The device is connected")
        print("   2. Another program is using the port")
        print("   3. You have permission to access the port")
        print("   4. The correct COM port is specified")
    except PermissionError as e:
        print(f"❌ Permission denied for {SERIAL_PORT}: {e}")
        print("   Try closing other applications that might be using this port")
    except Exception as e:
        print(f"⚠️ Unexpected serial error: {e}")

async def main():
    global log_rows, clicked_frames

    # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # output_file = log_dir / f"touch_log_{timestamp}.xlsx"
    # clicked_file = log_dir / f"clicked_log_{timestamp}.xlsx"

    # Load participant info
    participant_info_file = Path(__file__).parent/ "participant_info.txt"
    if participant_info_file.exists():
      info = participant_info_file.read_text().strip().split(',')
      if len(info) >= 5:
        participant_name, conditions, attempts, ID, delaytime = info
        attempts = int(attempts)
        ID = int(ID)
        delaytime = int(delaytime)
      else:
        participant_name = "default"
        conditions = "unknown"
        attempts = 0
        ID = 0
        delaytime = 0
    else:
     participant_name = "default"
     conditions = "unknown"
     attempts = 0
     ID = 0
     delaytime = 0

    # output_file = log_dir / f"{base_filename}_touch_log.xlsx"
    # clicked_file = log_dir / f"{base_filename}_clicked_log.xlsx"

    output_file = log_dir / f"{participant_name}_{conditions}_ID{ID}_{attempts}_{delaytime}_touch_log.xlsx"
    clicked_file = log_dir / f"{participant_name}_{conditions}_ID{ID}_{attempts}_{delaytime}_clicked_log.xlsx"

    # output_file = log_dir / f"touch_log.xlsx"
    # clicked_file = log_dir / f"clicked_log.xlsx"


    log_rows.clear()
    clicked_frames.clear()

    serial_thread = Thread(target=listen_serial, daemon=True)
    serial_thread.start()

    connection = await qtm_rt.connect(QTM_HOST)
    await connection.stream_frames(components=['3d', '6d'], on_packet=handle_qtm_data)

    print("🟢 Logging started...")

    # Wait for 7 clicks, then stop
    while len(clicked_frames) < 7:
        await asyncio.sleep(0.05)
        streaming_enabled = False

    print("🛑 Logging complete. Saving data...")

    # Save Excel
    wb_all = Workbook()
    ws_all = wb_all.active
    ws_all.title = "Touch Log"

    wb_clicked = Workbook()
    ws_clicked = wb_clicked.active
    ws_clicked.title = "Clicked Touches"

    headers = ['Frame', 'Pen X', 'Pen Y', 'Pen Z', 'Distance to Plane (mm)', 'Local X', 'Local Y', 'Touch Status', 'Clicked']
    ws_all.append(headers)
    ws_clicked.append(headers)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF475", end_color="FFF475", fill_type="solid")
    bold_font = Font(bold=True)

    # log_rows = set(log_rows)

    for row in log_rows:
            frame = row[0]
            clicked = 1 if frame in clicked_frames else 0
            full_row = row + [clicked]

            # check if full_row already exists in ws_all
            # if any(existing_row[:7] == full_row[:7] for existing_row in ws_all.iter_rows(min_row=2, values_only=True)):
            #     continue  # Skip duplicates
            
            ws_all.append(full_row)

            last_row = ws_all.max_row
            status_cell = ws_all.cell(row=last_row, column=8)
            if "Green" in row[7]:
                status_cell.fill = green_fill
            elif "Red" in row[7]:
                status_cell.fill = red_fill
            if clicked:
                for col in range(1, len(full_row) + 1):
                    cell = ws_all.cell(row=last_row, column=col)
                    cell.fill = yellow_fill
                    cell.font = bold_font
                ws_clicked.append(full_row)

    wb_all.save(output_file)
    wb_clicked.save(clicked_file)
    print(f"✅ Data saved to: {output_file}")
    print(f"✅ Clicked data saved to: {clicked_file}")
    time.sleep(2)
    sys.exit(0)

        # start_signal.unlink(missing_ok=True)
        # stop_signal.unlink(missing_ok=True)

if __name__ == "__main__":
    asyncio.run(main())
