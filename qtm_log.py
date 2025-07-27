import time
import asyncio
from datetime import datetime
from threading import Thread
import serial
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import qtm_rt
import os

# --- Configuration ---
QTM_HOST = '127.0.0.1'
DURATION_SECONDS = 1000
SERIAL_PORT = 'COM7'
BAUD_RATE = 9600

# --- Wait for participant folder path ---
session_file = Path(__file__).parent / "current_session_path.txt"
print("⏳ Waiting for participant folder info...")

max_wait = 15  # seconds
start = time.time()
while not session_file.exists() and (time.time() - start < max_wait):
    time.sleep(0.5)

if not session_file.exists():
    print("❌ Session path not found in time.")
    exit(1)

with open(session_file, "r") as f:
    folder_path = f.read().strip()

log_dir = Path(folder_path)
log_dir.mkdir(parents=True, exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = log_dir / f"touch_log_{timestamp}.xlsx"
CLICKED_ONLY_FILE = log_dir / f"clicked_log_{timestamp}.xlsx"

print(f"📁 Output will be saved to: {OUTPUT_FILE}")
print(f"📁 Clicked rows saved to: {CLICKED_ONLY_FILE}")

# --- Globals ---
log_rows = []
clicked_frames = set()
latest_frame = None
latest_qtm = None

def rect_point_to_local_xy(corners, point):
    corners = np.array(corners)
    point = np.array(point)
    center = np.mean(corners, axis=0)
    u = corners[1] - corners[0]
    v = corners[3] - corners[0]
    u_norm = u / np.linalg.norm(u)
    v_norm = v / np.linalg.norm(v)
    vec = point - center
    x_local = np.dot(vec, u_norm)
    y_local = np.dot(vec, v_norm)
    return x_local, y_local

def distance_point_to_plane(point, plane_point, plane_normal):
    point = np.array(point)
    plane_point = np.array(plane_point)
    plane_normal = np.array(plane_normal)
    plane_normal = plane_normal / np.linalg.norm(plane_normal)
    return np.dot(point - plane_point, plane_normal)

def handle_qtm_data(packet):
    global latest_frame, latest_qtm

    try:
        frame = packet.framenumber
        latest_frame = frame
        latest_qtm = []

        header, markers = packet.get_3d_markers()
        marker_xyz = [[m.x, m.y, m.z] for m in markers]
        if len(marker_xyz) >= 8:
            screen_corners = marker_xyz[4:8]
            pen_tip = marker_xyz[3]

            v1 = np.array(screen_corners[1]) - np.array(screen_corners[0])
            v2 = np.array(screen_corners[3]) - np.array(screen_corners[0])
            normal = np.cross(v1, v2)
            normal /= np.linalg.norm(normal)

            dist = abs(distance_point_to_plane(pen_tip, screen_corners[0], normal))
            status = "Not Touching"
            x_local = y_local = 'NaN'

            if dist < 6.0:
                x_local, y_local = rect_point_to_local_xy(screen_corners, pen_tip)
                width = np.linalg.norm(np.array(screen_corners[1]) - np.array(screen_corners[0]))
                height = np.linalg.norm(np.array(screen_corners[3]) - np.array(screen_corners[0]))
                if -width/2 <= x_local <= width/2 and -height/2 <= y_local <= height/2:
                    status = "Touching (Green)"
                else:
                    status = "Outside Bounds (Red)"

            log_rows.append([
                frame, pen_tip[0], pen_tip[1], pen_tip[2],
                dist, x_local, y_local, status
            ])
    except Exception as e:
        print(f"❌ QTM error: {e}")

async def start_qtm_listener():
    conn = await qtm_rt.connect(QTM_HOST)
    await conn.stream_frames(components=['3d', '6d'], on_packet=handle_qtm_data)

async def wait_for_valid_qtm():
    while True:
        await asyncio.sleep(0.01)
        if latest_qtm is not None:
            break

def listen_serial():
    try:
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=0.1)
        print("✅ Listening to serial...")
        while True:
            line = ser.readline().decode().strip()
            if line == '1' and latest_frame is not None:
                clicked_frames.add(latest_frame)
                print(f"🔘 Button clicked at frame: {latest_frame}")
    except Exception as e:
        print("⚠️ Serial error:", e)

def save_to_excel():
    wb_all = Workbook()
    ws_all = wb_all.active
    ws_all.title = "Touch Log"

    wb_clicked = Workbook()
    ws_clicked = wb_clicked.active
    ws_clicked.title = "Clicked Touches"

    headers = ['Frame', 'Pen X', 'Pen Y', 'Pen Z', 'Distance to Plane (mm)',
               'Local X', 'Local Y', 'Touch Status', 'Clicked']
    ws_all.append(headers)
    ws_clicked.append(headers)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF475", end_color="FFF475", fill_type="solid")
    bold_font = Font(bold=True)

    for row in log_rows:
        frame = row[0]
        clicked = 1 if frame in clicked_frames else 0
        full_row = row + [clicked]
        ws_all.append(full_row)

        last_row = ws_all.max_row
        status_cell = ws_all.cell(row=last_row, column=8)
        if "Green" in row[7]:
            status_cell.fill = green_fill
        elif "Red" in row[7]:
            status_cell.fill = red_fill
        if clicked:
            for col in range(1, len(full_row) + 1):
                cell = ws_all.cell(row=last_row, column=col)
                cell.fill = yellow_fill
                cell.font = bold_font
            ws_clicked.append(full_row)

    wb_all.save(OUTPUT_FILE)
    wb_clicked.save(CLICKED_ONLY_FILE)
    print(f"✅ Data saved to: {OUTPUT_FILE}")
    print(f"✅ Clicked data saved to: {CLICKED_ONLY_FILE}")

async def main():
    serial_thread = Thread(target=listen_serial, daemon=True)
    serial_thread.start()

    asyncio.ensure_future(start_qtm_listener())
    await wait_for_valid_qtm()
    start_time = time.time()

    try:
        while (time.time() - start_time) < DURATION_SECONDS:
            await asyncio.sleep(0.01)
    finally:
        save_to_excel()

if __name__ == '__main__':
    time.sleep(2)
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("🛑 Interrupted.")
        save_to_excel()
